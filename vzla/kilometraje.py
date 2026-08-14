# ==========================================
# Archivo: kilometraje.py
# Modulo: Construccion de Kilometraje
# - Lee cualquier Excel (incl. .xlsm)
# - Plantilla desde Google Sheet (PLANTILLA KM)
# - Calcula AYER / RECORRIDO / HOY y TOMAR = HOY - AYER (automatico)
# - Edicion en la app + descarga + empuje a KM-TABLERO / KM-ODOMETRO
# ==========================================
import streamlit as st
import gspread
from google.oauth2.service_account import Credentials
import textwrap
import re
import io
import pandas as pd
from datetime import datetime, date
import streamlit.components.v1 as components

SHEET_ID = "1D7w0ABnnatGd83TpJHFeVxxLOKRqoBBFbM9FyYYdFEg"
COLS_PLANTILLA = ["CLASIFICACION", "UNIDAD", "TIPO", "AYER", "HOY", "RECORRIDO", "TOMAR"]

# ==========================================
# CONEXION
# ==========================================
@st.cache_resource
def cliente_sheets():
    cred = dict(st.secrets["gcp_service_account"])
    lk = cred["private_key"]
    lk = (lk.replace("-----BEGIN PRIVATE KEY-----", "").replace("-----END PRIVATE KEY-----", "")
            .replace("\\n", "").replace("\n", "").replace(" ", ""))
    cred["private_key"] = "-----BEGIN PRIVATE KEY-----\n" + "\n".join(textwrap.wrap(lk, 64)) + "\n-----END PRIVATE KEY-----\n"
    sc = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']
    return gspread.authorize(Credentials.from_service_account_info(cred, scopes=sc))

def abrir_libro():
    return cliente_sheets().open_by_key(SHEET_ID)

def buscar_ws(libro, *claves):
    """Encuentra una hoja cuyo título contenga TODAS las claves (sin importar acentos/espacios)."""
    def n(x): return re.sub(r'[^A-Z0-9]', '', str(x).upper())
    claves_n = [n(k) for k in claves]
    for w in libro.worksheets():
        t = n(w.title)
        if all(k in t for k in claves_n):
            return w
    return None

# ==========================================
# NUMEROS
# ==========================================
def to_num(x):
    """'54.773,54' -> 54773.54 ; '484,00' -> 484.0 ; '1560' -> 1560.0"""
    s = str(x).replace("Kms", "").strip()
    if s == "" or s.lower() == "nan":
        return None
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None

def fmt_km(v):
    """480352.5 -> '480.352,50'. Vacío si None/NaN."""
    if v is None:
        return ""
    try:
        if pd.isna(v):
            return ""
    except Exception:
        pass
    return f"{float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def plate(x):
    return str(x).strip().upper()

# ==========================================
# LECTURA DE EXCEL (cualquiera, incl .xlsm)
# ==========================================
def _leer_excel(file, as_str=True):
    nombre = getattr(file, "name", "").lower()
    dt = str if as_str else None
    if nombre.endswith(".xls"):
        try:
            return pd.read_excel(file, header=None, dtype=dt, engine="xlrd")
        except Exception:
            file.seek(0)
            return pd.read_html(file)[0]
    # xlsx y xlsm -> openpyxl
    return pd.read_excel(file, header=None, dtype=dt, engine="openpyxl")

def _parse_fecha(v):
    """Interpreta una celda de encabezado como fecha: fecha real, serial de Excel, texto o 'dd-mes'."""
    if v is None:
        return None
    if isinstance(v, (pd.Timestamp, datetime)):
        return v.date()
    if isinstance(v, date):
        return v
    # número serie de Excel
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            if 20000 <= float(v) <= 80000:
                return pd.to_datetime(float(v), unit="D", origin="1899-12-30").date()
        except Exception:
            pass
    s = str(v).strip()
    if not s or s.lower() == "nan":
        return None
    if re.fullmatch(r'\d{4,6}(\.0)?', s):  # serial como texto
        try:
            return pd.to_datetime(float(s), unit="D", origin="1899-12-30").date()
        except Exception:
            pass
    iso = bool(re.match(r'^\d{4}[-/]\d{1,2}[-/]\d{1,2}', s))  # año primero -> no dayfirst
    d = pd.to_datetime(s, dayfirst=not iso, errors="coerce")
    if pd.notna(d):
        return d.date()
    meses = {'ene': 1, 'feb': 2, 'mar': 3, 'abr': 4, 'may': 5, 'jun': 6,
             'jul': 7, 'ago': 8, 'sep': 9, 'set': 9, 'oct': 10, 'nov': 11, 'dic': 12}
    m = re.search(r'(\d{1,2})[\s\-/]*([a-zA-Zñ]{3,})', s.lower())
    if m:
        dia = int(m.group(1)); mes = meses.get(m.group(2)[:3])
        if mes:
            am = re.search(r'(20\d{2})', s)
            anio = int(am.group(1)) if am else datetime.now().year
            try:
                return date(anio, mes, dia)
            except Exception:
                pass
    return None

def leer_odometro(file):
    """Devuelve dict UNIDAD -> recorrido (km del odómetro GPS)."""
    raw = _leer_excel(file)
    hrow = None
    for r in range(min(30, len(raw))):
        fila = [str(x).upper() for x in raw.iloc[r].tolist()]
        if any("UNIDAD" in x for x in fila) and any("ODÓMETRO" in x or "ODOMETRO" in x for x in fila):
            hrow = r
            break
    if hrow is None:
        hrow = 8
    df = raw.iloc[hrow:].reset_index(drop=True)
    df.columns = [str(c).strip() for c in df.iloc[0].tolist()]
    df = df.iloc[1:].reset_index(drop=True)
    col_u = next((c for c in df.columns if "UNIDAD" in str(c).upper()), None)
    col_o = next((c for c in df.columns if "ODÓMETRO" in str(c).upper() or "ODOMETRO" in str(c).upper()), None)
    if not col_u or not col_o:
        return {}
    mp = {}
    for _, r in df.iterrows():
        u = plate(r[col_u])
        if u and u != "NAN":
            mp[u] = to_num(r[col_o]) or 0.0
    return mp

def leer_km_diario(file, fecha_sel):
    """Devuelve (dict UNIDAD -> km de la fecha, etiqueta_columna) del archivo de km diario.
    Detecta la fecha aunque venga como fecha real, serial de Excel o texto."""
    raw = _leer_excel(file, as_str=False)
    hrow = None
    for r in range(min(15, len(raw))):
        fila = [str(x).strip().upper() for x in raw.iloc[r].tolist()]
        if "UNIDAD" in fila:
            hrow = r
            break
    if hrow is None:
        hrow = 0
    headers = raw.iloc[hrow].tolist()
    data = raw.iloc[hrow + 1:].reset_index(drop=True)

    col_u_idx = next((i for i, h in enumerate(headers) if str(h).strip().upper() == "UNIDAD"), None)
    col_f_idx, etiqueta = None, None
    for i, h in enumerate(headers):
        if _parse_fecha(h) == fecha_sel:
            col_f_idx, etiqueta = i, str(h)
            break
    if col_u_idx is None or col_f_idx is None:
        return {}, None

    mp = {}
    for _, row in data.iterrows():
        u = plate(row.iloc[col_u_idx])
        if u and u != "NAN":
            mp[u] = to_num(row.iloc[col_f_idx])
    return mp, etiqueta

def leer_plantilla_sheet(libro):
    ws = buscar_ws(libro, "PLANTIL", "KM")
    if ws is None:
        return None, "No encontré la hoja PLANTILLA KM."
    df = pd.DataFrame(ws.get_all_records())
    if df.empty:
        return None, "La hoja PLANTILLA KM está vacía."
    df.columns = [str(c).strip().upper() for c in df.columns]
    for c in COLS_PLANTILLA:
        if c not in df.columns:
            df[c] = ""
    df = df[COLS_PLANTILLA]
    df = df[df["UNIDAD"].astype(str).str.strip() != ""]
    return df.reset_index(drop=True), "OK"

# ==========================================
# PROCESO
# ==========================================
def procesar(plantilla, odo_map, ayer_map):
    df = plantilla.copy()
    df["UNIDAD"] = df["UNIDAD"].astype(str).str.strip()
    key = df["UNIDAD"].apply(plate)
    df["AYER"] = key.map(ayer_map)
    df["RECORRIDO"] = key.map(odo_map).fillna(0.0)
    df["AYER"] = pd.to_numeric(df["AYER"], errors="coerce")
    df["RECORRIDO"] = pd.to_numeric(df["RECORRIDO"], errors="coerce").fillna(0.0)
    df["HOY"] = df["AYER"].fillna(0.0) + df["RECORRIDO"]
    # Regla: recorrido > 125 y CLASIFICACION != ODOMETRO -> HOY vacío (para llenar manual)
    clasif = df["CLASIFICACION"].fillna("").astype(str).str.upper().str.strip()
    cond = (df["RECORRIDO"] > 125) & (clasif != "ODOMETRO")
    df.loc[cond, "HOY"] = None
    df["TOMAR"] = pd.to_numeric(df["HOY"], errors="coerce") - pd.to_numeric(df["AYER"], errors="coerce")
    # Columnas numéricas float: los vacíos quedan como NaN (el editor los muestra en blanco, no 'None')
    for c in ["AYER", "HOY", "RECORRIDO", "TOMAR"]:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    return df

def recalcular_tomar(df):
    df = df.copy()
    df["TOMAR"] = pd.to_numeric(df["HOY"], errors="coerce") - pd.to_numeric(df["AYER"], errors="coerce")
    return df

# ==========================================
# EXPORTAR PLANTILLA LLENA (Excel)
# ==========================================
def exportar_excel(df):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font
    wb = Workbook(); ws = wb.active; ws.title = "PLANTILLA"
    ws.append(COLS_PLANTILLA)
    for c in ws[1]:
        c.font = Font(bold=True)
    azul = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    rojo = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    for _, r in df.iterrows():
        ayer = pd.to_numeric(pd.Series([r["AYER"]]), errors="coerce").iloc[0]
        hoy = pd.to_numeric(pd.Series([r["HOY"]]), errors="coerce").iloc[0]
        rec = pd.to_numeric(pd.Series([r["RECORRIDO"]]), errors="coerce").iloc[0]
        tom = pd.to_numeric(pd.Series([r["TOMAR"]]), errors="coerce").iloc[0]
        fila = [r.get("CLASIFICACION", ""), r.get("UNIDAD", ""), r.get("TIPO", ""),
                fmt_km(ayer), fmt_km(hoy), fmt_km(rec), fmt_km(tom)]
        ws.append(fila)
        rid = ws.max_row
        if pd.notna(ayer) and pd.notna(hoy) and ayer == hoy and ayer != 0:
            ws.cell(row=rid, column=4).fill = azul
            ws.cell(row=rid, column=5).fill = azul
        if pd.notna(rec) and rec > 1000:
            ws.cell(row=rid, column=6).fill = rojo
    ws.auto_filter.ref = f"A1:G1"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()

# ==========================================
# EMPUJE A KM-TABLERO / KM-ODOMETRO
# ==========================================
def _fecha_col_index(ws, fecha_sel):
    """Busca en la fila 1 la columna cuya fecha coincide. Devuelve (indice_1based | None, header)."""
    header = ws.row_values(1)
    for i, h in enumerate(header):
        if _parse_fecha(h) == fecha_sel:
            return i + 1, header
    return None, header

def _col_unidad(header):
    for i, h in enumerate(header):
        if str(h).strip().upper() in ("UNIDAD", "PLACA"):
            return i + 1
    return 2  # por defecto columna B (A=CLASIFICACION, B=UNIDAD, C=TIPO)

def empujar_a_hoja(ws, fecha_sel, valores_por_placa):
    """Escribe valores (dict UNIDAD->valor) en la columna de la fecha, formato '54.888,90 Kms'.
    La placa se busca en la columna UNIDAD (por defecto B). Si la fecha no existe, crea la columna al final."""
    header = ws.row_values(1)
    col_u = _col_unidad(header)
    col_fecha, header = _fecha_col_index(ws, fecha_sel)
    if col_fecha is None:
        col_fecha = len(header) + 1
        if hasattr(fecha_sel, "day"):
            etiqueta = f"{fecha_sel.day}/{fecha_sel.month}/{fecha_sel.year}"
        else:
            etiqueta = str(fecha_sel)
        ws.update_cell(1, col_fecha, etiqueta)
    # mapa placa -> fila (desde la fila 2)
    placas_col = ws.col_values(col_u)
    fila_de = {}
    for idx, p in enumerate(placas_col[1:], start=2):
        if str(p).strip():
            fila_de[plate(p)] = idx
    from gspread.utils import rowcol_to_a1
    updates = []
    for u, val in valores_por_placa.items():
        if val is None or (isinstance(val, float) and pd.isna(val)):
            continue
        f = fila_de.get(plate(u))
        if f is None:
            continue
        texto = fmt_km(val) + " Kms"   # respeta tu formato: 54.888,90 Kms
        updates.append({"range": rowcol_to_a1(f, col_fecha), "values": [[texto]]})
    if updates:
        ws.batch_update(updates, value_input_option="USER_ENTERED")
    return len(updates)

# ==========================================
# INTERFAZ
# ==========================================
st.title("🛞 Construcción de Kilometraje")

try:
    libro = abrir_libro()
except Exception as e:
    st.error("No se pudo conectar al Google Sheet.")
    st.info("Comparte el Sheet como Editor con `bot-pizarra@pcd-drotaca.iam.gserviceaccount.com`.")
    st.exception(e); st.stop()

st.markdown("#### 1) Archivos de origen")
c1, c2 = st.columns(2)
f_odo = c1.file_uploader("📂 Odómetro (GPS)", type=["xlsx", "xls", "xlsm"], key="km_odo")
f_km = c2.file_uploader("📂 Kilometraje diario (para el 'AYER')", type=["xlsx", "xls", "xlsm"], key="km_dia")
fecha_ayer = st.date_input("📅 Fecha del 'kilometraje de ayer' (columna a tomar del diario)",
                           value=date.today(), format="DD/MM/YYYY")

if st.button("⚙️ Procesar (traer plantilla del Sheet y calcular)", type="primary", use_container_width=True):
    if not f_odo or not f_km:
        st.warning("Sube ambos archivos: Odómetro y Kilometraje diario.")
    else:
        try:
            plantilla, msg = leer_plantilla_sheet(libro)
            if plantilla is None:
                st.error(msg); st.stop()
            odo_map = leer_odometro(f_odo)
            ayer_map, col_km = leer_km_diario(f_km, fecha_ayer)
            if not col_km:
                st.error(f"No encontré la columna de la fecha {fecha_ayer.strftime('%d/%m/%Y')} en el archivo de km diario.")
                st.stop()
            df = procesar(plantilla, odo_map, ayer_map)
            st.session_state["km_df"] = df
            st.session_state["km_fecha"] = fecha_ayer
            st.success(f"✅ Procesado. AYER tomado de la columna '{col_km}'. "
                       f"Odómetro con {len(odo_map)} unidades.")
        except Exception as e:
            st.error(f"Error al procesar: {e}")
            st.exception(e)

if "km_df" in st.session_state:
    df = st.session_state["km_df"]

    # Alertas
    rec_alto = df[pd.to_numeric(df["RECORRIDO"], errors="coerce") > 1000]["UNIDAD"].tolist()
    faltan_hoy = df[pd.to_numeric(df["HOY"], errors="coerce").isna()]["UNIDAD"].tolist()
    if rec_alto:
        st.warning("⚠️ Recorrido > 1000 km (revisar): " + ", ".join(rec_alto))
    if faltan_hoy:
        st.info(f"✏️ {len(faltan_hoy)} unidades quedaron con HOY vacío para que las completes manualmente "
                "(recorrido > 125 y no son ODOMETRO): " + ", ".join(faltan_hoy))

    st.markdown("#### 2) Edita el HOY que falte (los demás se calculan solos)")
    st.caption("Edita AYER/HOY/RECORRIDO si hace falta. TOMAR = HOY − AYER se recalcula al pulsar el botón de abajo.")
    edit = st.data_editor(
        df, use_container_width=True, hide_index=True, num_rows="dynamic",
        column_config={
            "AYER": st.column_config.NumberColumn("AYER", format="%.2f"),
            "HOY": st.column_config.NumberColumn("HOY", format="%.2f"),
            "RECORRIDO": st.column_config.NumberColumn("RECORRIDO", format="%.2f"),
            "TOMAR": st.column_config.NumberColumn("TOMAR (auto)", format="%.2f", disabled=True),
        }, key="km_editor"
    )

    if st.button("🔄 Recalcular TOMAR (HOY − AYER)", use_container_width=True):
        st.session_state["km_df"] = recalcular_tomar(edit)
        st.rerun()

    dfx = recalcular_tomar(edit)

    st.markdown("#### 3) Vista con formato (480.352,50)")
    vista = pd.DataFrame({
        "CLASIFICACION": dfx["CLASIFICACION"],
        "UNIDAD": dfx["UNIDAD"],
        "TIPO": dfx["TIPO"],
        "AYER": dfx["AYER"].apply(lambda v: fmt_km(to_num(v) if isinstance(v, str) else v)),
        "HOY": dfx["HOY"].apply(lambda v: fmt_km(to_num(v) if isinstance(v, str) else v)),
        "RECORRIDO": dfx["RECORRIDO"].apply(lambda v: fmt_km(to_num(v) if isinstance(v, str) else v)),
        "TOMAR": dfx["TOMAR"].apply(lambda v: fmt_km(to_num(v) if isinstance(v, str) else v)),
    })
    st.dataframe(vista, use_container_width=True, hide_index=True)

    st.download_button("⬇️ Descargar Plantilla KM (Excel)", data=exportar_excel(dfx),
                       file_name="Plantilla_KM_llena.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       use_container_width=True)

    st.markdown("---")
    st.markdown("#### 4) Cargar al Excel master (Google Sheet)")
    fecha_push = st.date_input("📅 Fecha destino en KM-TABLERO / KM-ODOMETRO",
                               value=st.session_state.get("km_fecha", date.today()), format="DD/MM/YYYY", key="fpush")
    st.caption("Escribe la columna HOY en la hoja **KM-TABLERO** y la columna TOMAR en la hoja **KM-ODOMETRO**, "
               "en la columna de esa fecha (si no existe, la crea al final).")
    confirmar = st.checkbox("Confirmo que quiero escribir en el Google Sheet")
    if st.button("💾 Cargar datos al master", type="primary", use_container_width=True, disabled=not confirmar):
        try:
            dff = recalcular_tomar(edit)
            hoy_map = {plate(r["UNIDAD"]): to_num(r["HOY"]) if isinstance(r["HOY"], str) else r["HOY"] for _, r in dff.iterrows()}
            tomar_map = {plate(r["UNIDAD"]): to_num(r["TOMAR"]) if isinstance(r["TOMAR"], str) else r["TOMAR"] for _, r in dff.iterrows()}
            ws_tab = buscar_ws(libro, "TABLERO")
            ws_odo = buscar_ws(libro, "ODOMETRO")
            if ws_tab is None or ws_odo is None:
                st.error("No encontré las hojas KM-TABLERO y/o KM-ODOMETRO.")
            else:
                n1 = empujar_a_hoja(ws_tab, fecha_push, hoy_map)
                n2 = empujar_a_hoja(ws_odo, fecha_push, tomar_map)
                st.cache_data.clear()
                st.success(f"✅ Cargado. KM-TABLERO: {n1} valores (HOY) · KM-ODOMETRO: {n2} valores (TOMAR).")
        except Exception as e:
            st.error(f"No se pudo cargar al master: {e}")
            st.exception(e)
else:
    st.info("Sube los archivos y pulsa Procesar.")
