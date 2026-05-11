import streamlit as st
import pandas as pd
from datetime import datetime
import gspread
from google.oauth2.service_account import Credentials
import textwrap
import openpyxl
import io
import os

# --- 1. CONFIGURACIÓN DE CONEXIÓN ---
CREDENCIALES_GOOGLE = dict(st.secrets["gcp_service_account"])
llave_sucia = CREDENCIALES_GOOGLE["private_key"]
llave_limpia = llave_sucia.replace("-----BEGIN PRIVATE KEY-----", "").replace("-----END PRIVATE KEY-----", "").replace("\\n", "").replace("\n", "").replace(" ", "")
llave_perfecta = "-----BEGIN PRIVATE KEY-----\n" + "\n".join(textwrap.wrap(llave_limpia, 64)) + "\n-----END PRIVATE KEY-----\n"
CREDENCIALES_GOOGLE["private_key"] = llave_perfecta

def obtener_cliente_sheets():
    alcance = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']
    credenciales = Credentials.from_service_account_info(CREDENCIALES_GOOGLE, scopes=alcance)
    return gspread.authorize(credenciales)

def conectar_bd_compras():
    return obtener_cliente_sheets().open_by_key("1M9VQOHU6LHniSBu3A_o2qDFlZITrDRm_f7qYGU5iOX8")

# --- 2. CARGA INTELIGENTE Y AUTO-REPARACIÓN DE DATOS ---
try:
    doc = conectar_bd_compras()
    hoja_sol = doc.worksheet("SOLICITUDES")
    headers = [str(h).strip() for h in hoja_sol.row_values(1)]
    
    # Auto-completar columnas faltantes en tu Google Sheet
    columnas_requeridas = ["Fecha_Entrega", "Dias_Resolucion", "Nota"]
    for col in columnas_requeridas:
        if col not in headers:
            headers.append(col)
            hoja_sol.update_cell(1, headers.index(col) + 1, col)
            
    data_raw = hoja_sol.get_all_records()
    df_global = pd.DataFrame(data_raw)
    
    if not df_global.empty:
        df_global.columns = [str(c).strip() for c in df_global.columns]
        # Normalizamos nombres por si tu sheet dice "Fecha" en lugar de "Fecha_Solicitud"
        rename_map = {'Fecha': 'Fecha_Solicitud', 'Tipo_Solicitud': 'Categoria'}
        df_global.rename(columns=rename_map, inplace=True)
        
        # Garantizar que las columnas existan en memoria para no dar KeyError
        for col_req in ['Fecha_Entrega', 'Dias_Resolucion', 'Nota', 'Estatus', 'Fecha_Solicitud']:
            if col_req not in df_global.columns:
                df_global[col_req] = ""
except Exception as e:
    df_global = pd.DataFrame()
    headers = []

st.title("🛒 Gestión y Cronometría de Compras")
tab_nueva, tab_control, tab_metricas = st.tabs(["📝 Crear Solicitud", "✅ Check de Entrega", "📊 Historial y Auditoría"])

# ---------------------------------------------------------
# PESTAÑA 1: GENERAR SOLICITUD
# ---------------------------------------------------------
with tab_nueva:
    with st.container(border=True):
        col_id, col_cat = st.columns(2)
        with col_id:
            try:
                hoja_cnt = doc.worksheet("CONTADOR")
                proximo_id = int(hoja_cnt.acell('A2').value)
            except: proximo_id = 1
            st.subheader(f"Planilla N° {proximo_id}")
            fecha_sol = st.date_input("Fecha Solicitud:", datetime.now())
        with col_cat:
            categoria = st.selectbox("Tipo de Solicitud:", [
                "Solicitud para stock (almacén de flota)", 
                "Solicitud para vehículos"
            ])

    if 'filas_compras' not in st.session_state:
        st.session_state['filas_compras'] = pd.DataFrame([{"Cantidad": 1, "Descripción": ""}] * 5)
    if 'archivo_excel_listo' not in st.session_state:
        st.session_state['archivo_excel_listo'] = None
    if 'nombre_excel_listo' not in st.session_state:
        st.session_state['nombre_excel_listo'] = None
    if 'mensaje_exito' not in st.session_state:
        st.session_state['mensaje_exito'] = None

    df_editado = st.data_editor(st.session_state['filas_compras'], num_rows="dynamic", use_container_width=True, hide_index=True)
    nota_p = st.text_area("Notas para el Excel (B22):")

    col_btn1, col_btn2 = st.columns(2)
    with col_btn1:
        if st.button("🔄 LIMPIAR / REGRESAR", use_container_width=True):
            st.session_state['filas_compras'] = pd.DataFrame([{"Cantidad": 1, "Descripción": ""}] * 5)
            st.session_state['archivo_excel_listo'] = None
            st.session_state['mensaje_exito'] = None
            st.rerun()

    with col_btn2:
        if st.button("💾 GUARDAR SOLICITUD", type="primary", use_container_width=True):
            df_valido = df_editado[df_editado['Descripción'].str.strip() != ""].copy()
            if df_valido.empty:
                st.error("Debes llenar al menos un ítem con descripción.")
            else:
                with st.spinner("Alineando columnas y guardando..."):
                    registros = []
                    for i, (_, r) in enumerate(df_valido.iterrows(), start=1):
                        # Diccionario con todos los posibles nombres que tengas en tu Sheet
                        fila_dict = {
                            "ID_Planilla": proximo_id,
                            "Fecha": fecha_sol.strftime("%d/%m/%Y"),
                            "Fecha_Solicitud": fecha_sol.strftime("%d/%m/%Y"),
                            "Tipo_Solicitud": categoria,
                            "Categoria": categoria,
                            "Item_No": i,
                            "Cantidad": r['Cantidad'],
                            "Descripcion": r['Descripción'].upper(),
                            "Estatus": "PENDIENTE",
                            "Fecha_Entrega": "",
                            "Dias_Resolucion": "",
                            "Usuario": st.session_state.get('usuario','admin_vzla'),
                            "Nota": nota_p
                        }
                        # Acomodar exactamente en el orden de los headers
                        fila_ordenada = [fila_dict.get(h, "") for h in headers]
                        registros.append(fila_ordenada)
                    
                    hoja_sol.append_rows(registros)
                    doc.worksheet("CONTADOR").update_acell('A2', proximo_id + 1)
                    
                    # Generar Excel
                    ruta_excel = os.path.join(os.path.dirname(__file__), "Planilla de Solicitud de compra.xlsx")
                    if os.path.exists(ruta_excel):
                        wb = openpyxl.load_workbook(ruta_excel)
                        ws = wb.active
                        ws["C6"] = proximo_id
                        ws["B22"] = f"NOTA: {nota_p}"
                        for i, (idx, row) in enumerate(df_valido.iterrows(), start=8):
                            if i <= 20:
                                ws[f"C{i}"], ws[f"D{i}"] = row['Cantidad'], row['Descripción'].upper()
                        out = io.BytesIO()
                        wb.save(out)
                        
                        st.session_state['archivo_excel_listo'] = out.getvalue()
                        st.session_state['nombre_excel_listo'] = f"Solicitud_{proximo_id}.xlsx"
                        st.session_state['mensaje_exito'] = f"✅ ¡Planilla N° {proximo_id} guardada con éxito!"
                    else:
                        st.session_state['mensaje_exito'] = f"✅ Guardado. Falta molde Excel."

                    st.session_state['filas_compras'] = pd.DataFrame([{"Cantidad": 1, "Descripción": ""}] * 5)
                    st.rerun()

    if st.session_state['mensaje_exito']:
        st.success(st.session_state['mensaje_exito'])
    if st.session_state['archivo_excel_listo']:
        st.download_button(label="⬇️ DESCARGAR PLANILLA EXCEL OFICIAL", data=st.session_state['archivo_excel_listo'], file_name=st.session_state['nombre_excel_listo'], mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="secondary", use_container_width=True)

# ---------------------------------------------------------
# PESTAÑA 2: CHECK DE ENTREGA
# ---------------------------------------------------------
with tab_control:
    st.header("📋 Pendientes de Compra")
    if not df_global.empty and 'Estatus' in df_global.columns:
        df_p = df_global[df_global['Estatus'].astype(str).str.strip() == 'PENDIENTE'].copy()
        
        if df_p.empty:
            st.success("✅ No hay ítems pendientes.")
        else:
            st.dataframe(df_p[['ID_Planilla', 'Fecha_Solicitud', 'Descripcion', 'Cantidad']], use_container_width=True, hide_index=True)
            st.markdown("---")
            opciones = df_p.apply(lambda r: f"P-{r['ID_Planilla']} | {r['Cantidad']}x {r['Descripcion']}", axis=1).tolist()
            seleccion = st.selectbox("Seleccione el ítem recibido:", opciones)
            fecha_rec = st.date_input("Fecha de Recepción:", datetime.now())
            
            if st.button("🏁 CONFIRMAR ENTREGA", type="primary"):
                id_sel = int(seleccion.split("|")[0].replace("P-", "").strip())
                desc_sel = seleccion.split("|")[1].split("x ", 1)[1].strip()
                
                # Encontrar columnas dinámicamente
                col_estatus = headers.index("Estatus") + 1
                col_f_ent = headers.index("Fecha_Entrega") + 1
                col_dias = headers.index("Dias_Resolucion") + 1

                data_actual = hoja_sol.get_all_records()
                for idx, row in enumerate(data_actual):
                    if int(row['ID_Planilla']) == id_sel and str(row['Descripcion']).strip() == desc_sel:
                        fila_sheet = idx + 2
                        f_ini = datetime.strptime(str(row.get('Fecha', row.get('Fecha_Solicitud'))), "%d/%m/%Y")
                        f_fin = datetime.combine(fecha_rec, datetime.min.time())
                        dias = (f_fin - f_ini).days
                        
                        # Actualizar en las columnas correctas sin importar el orden
                        hoja_sol.update_cell(fila_sheet, col_estatus, "COMPRADO")
                        hoja_sol.update_cell(fila_sheet, col_f_ent, fecha_rec.strftime("%d/%m/%Y"))
                        hoja_sol.update_cell(fila_sheet, col_dias, dias)
                        
                        st.success(f"Entregado en {dias} días.")
                        st.session_state['archivo_excel_listo'] = None
                        st.session_state['mensaje_exito'] = None
                        st.rerun()
                        break

# ---------------------------------------------------------
# PESTAÑA 3: HISTORIAL Y AUDITORÍA
# ---------------------------------------------------------
with tab_metricas:
    st.header("📊 Auditoría de Cumplimiento y Eficiencia")
    
    if not df_global.empty and 'Estatus' in df_global.columns:
        df_audit = df_global.copy()
        df_audit['Estatus'] = df_audit['Estatus'].astype(str).str.strip()
        df_audit['Fecha_Solicitud'] = pd.to_datetime(df_audit['Fecha_Solicitud'], format='%d/%m/%Y', errors='coerce')
        
        df_comp = df_audit[df_audit['Estatus'] == 'COMPRADO'].copy()
        df_pend = df_audit[df_audit['Estatus'] == 'PENDIENTE'].copy()
        
        m1, m2, m3, m4 = st.columns(4)
        
        total_sol = len(df_audit)
        entregados = len(df_comp)
        pendientes = len(df_pend)
        efectividad = (entregados / total_sol * 100) if total_sol > 0 else 0
        
        promedio_dias = pd.to_numeric(df_comp['Dias_Resolucion'], errors='coerce').mean() if not df_comp.empty else 0

        m1.metric("Total Solicitudes", total_sol)
        m2.metric("Efectividad", f"{efectividad:.1f}%")
        m3.metric("Pendientes", pendientes, delta=f"Faltan {pendientes}", delta_color="inverse")
        m4.metric("Promedio Entrega", f"{promedio_dias:.1f} días")

        if pendientes >= 15:
            st.error(f"🚨 **ALERTA DE GESTIÓN:** Se han acumulado {pendientes} solicitudes sin atender. Es necesario revisar el flujo de compras.")

        st.markdown("---")
        st.subheader("📋 Detalle General de Movimientos")

        df_audit['DÍAS'] = 0
        hoy = datetime.now()
        for idx, row in df_audit.iterrows():
            if row['Estatus'] == 'PENDIENTE' and pd.notnull(row['Fecha_Solicitud']):
                df_audit.at[idx, 'DÍAS'] = (hoy - row['Fecha_Solicitud']).days
            else:
                df_audit.at[idx, 'DÍAS'] = pd.to_numeric(row.get('Dias_Resolucion', 0), errors='coerce')

        df_visual = df_audit[['Fecha_Solicitud', 'Descripcion', 'Cantidad', 'Usuario', 'Estatus', 'DÍAS']].copy()
        df_visual.columns = ['FECHA', 'ITEM / DESCRIPCIÓN', 'CANT', 'SOLICITADO POR', 'ESTADO', 'DÍAS']
        df_visual['FECHA'] = df_visual['FECHA'].dt.strftime('%d/%m/%Y')

        st.dataframe(
            df_visual.sort_values('FECHA', ascending=False),
            use_container_width=True, hide_index=True,
            column_config={
                "ESTADO": st.column_config.SelectboxColumn("ESTADO", options=["PENDIENTE", "COMPRADO"]),
                "DÍAS": st.column_config.NumberColumn("DÍAS", format="%d d")
            }
        )
    else:
        st.info("Aún no hay registros en el historial para mostrar métricas.")
